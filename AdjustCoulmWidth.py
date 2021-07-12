import openpyxl
from openpyxl.utils import get_column_letter

# YOUR EXCEL FILE WITH PATH, HERE I HAVE USED THE RELATIVE PATH
filename = f'SampleExcel.xlsx'

#Format the columns of excel file 
#Auto adjust column width       
wb = openpyxl.load_workbook(filename = f'{filename}')
for sheet in wb.sheetnames:       
    ws = wb[sheet]
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)  # Get the column name        
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1
        ws.column_dimensions[column].width = adjusted_width         
                                                    
    wb.save(f'{filename}')